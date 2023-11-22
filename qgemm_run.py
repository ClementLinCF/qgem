import subprocess
import re
import os
import time
import threading
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def execute_program(command):
    try:
        # Run qgemm.py with parameters and redirect stdout and stderr to qgemm_tmp.log
        with open('qgemm_tmp.log', 'w') as logfile:
            subprocess.run(command, stdout=logfile, stderr=subprocess.STDOUT, check=True)

    except subprocess.CalledProcessError as e:
        print(f"Error running qgemm.py: {e}")


def check_mnkabc_match(M, N, K, LDA, LDB, LDC):
    # Define the regular expression pattern to match the desired values
    pattern = r'-m\s+(\d+).*-n\s+(\d+).*-k\s+(\d+).*--lda\s+(\d+).*--ldb\s+(\d+).*--ldc\s+(\d+)'

    # Read from the log file
    with open('qgemm_tmp.log', 'r') as log_file:
        log_string = log_file.read()

    # Find all matches in the log string
    matches = re.findall(pattern, log_string)
        
    # os.remove('qgemm_tmp.log')

    if matches:
        for match in matches:
            m_value = int(match[0])
            n_value = int(match[1])
            k_value = int(match[2])
            lda_value = int(match[3])
            ldb_value = int(match[4])
            ldc_value = int(match[5])

            if (m_value == M and
                n_value == N and
                k_value == K and
                lda_value == LDA and
                ldb_value == LDB and
                ldc_value == LDC):
                print("****** m,n,k,lda,ldb,ldc are identical")
                return True
            else:
                print("****** m,n,k,lda,ldb,ldc are not identical *********")
                print(f"m_value: {m_value} vs self.M: {M}")
                print(f"n_value: {n_value} vs self.N: {N}")
                print(f"k_value: {k_value} vs self.K: {K}")
                print(f"lda_value: {lda_value} vs self.LDA: {LDA}")
                print(f"ldb_value: {ldb_value} vs self.LDB: {LDB}")
                print(f"ldc_value: {ldc_value} vs self.LDC: {LDC}")
                print("****************************************************")
                return False, m_value, n_value, k_value, lda_value, ldb_value, ldc_value
    else:
        print("**** Pattern not found in the log file.")
        return None


def get_TFLOPS():
    # Read from the log file
    with open('qgemm_tmp.log', 'r') as log_file:
        log_lines = log_file.readlines()

    # Define a regular expression pattern to match floating-point numbers
    pattern = r'\s+(\d+\.\d+)'

    # List to store the first numbers from each line that matches the pattern
    numbers = []

    # Iterate through each line in the log file
    for line in log_lines:
        match = re.search(pattern, line)
        if match:
            number = float(match.group(1))
            numbers.append(number)

    # Check if any numbers were found
    if numbers:
        # Find the maximum number from the list and divide it by 1000
        max_number = max(numbers) / 1000
        print(f"The maximum number divided by 1000 is: {max_number}")
        return max_number
        # Delete the log file after processing
        os.remove('qgemm_tmp.log')
    else:
        print("No matching numbers found in the log file.")

def produce_report(command, tflops, peak_tflops):
    column_names = ['PRECISION_A', 'PRECISION_B', 'PRECISION_C', 'COMPUTE_PRECISION',
                    'OP_A', 'OP_B', 'M', 'N', 'K', 'LDA', 'LDB', 'LDC', 'BATCH_COUNT',
                    'TIME_SPAN', 'EX', 'TFLOPS', 'PEAK_TFLOPS', 'efficiency', 'P_M', 'P_N', 'P_K', 'P_LDA', 'P_LDB', 'P_LDC', 'Status']

    # Load an existing Excel workbook if it exists, otherwise create a new one
    try:
        workbook = openpyxl.load_workbook('report.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet or create a new one if it doesn't exist
    if command[2] in workbook.sheetnames:
        sheet = workbook[command[2]]  # Use the existing sheet with the specified name
        next_row = sheet.max_row + 1  # Get the next available row
    else:
        sheet = workbook.create_sheet(title=command[2])  # Create a new sheet with the specified name
        next_row = 1  # Start from the first row

        # Write column names to the first row of the sheet
        for col, column_name in enumerate(column_names, start=1):
            cell = sheet.cell(row=next_row, column=col, value=column_name)
            cell.font = Font(bold=True)
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = max(len(str(column_name)) + 2, 10)  # Adjust column width
        next_row += 1  # Move to the next row for data

    # Write command data to the second row of the sheet
    for col, value in enumerate(command[2:], start=1):
        cell = sheet.cell(row=next_row, column=col, value=value)
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = max(len(str(value)) + 2, 10)  # Adjust column width

    # Write TFLOPS and PEAK_TFLOPS values to the second row of the sheet
    sheet.cell(row=next_row, column=len(command) + 1 - 2, value=tflops)
    sheet.cell(row=next_row, column=len(command) + 2 - 2, value=peak_tflops)

    # efficiency = tflops / peak_tflops * 100
    efficiency = 0
    if tflops != None and peak_tflops != 0:
        efficiency = tflops / peak_tflops
    # sheet.cell(row=next_row, column=len(command) + 3 - 2, value=efficiency)
    # sheet[f"{openpyxl.utils.get_column_letter(len(command) + 3 - 2)}2"].number_format = '0.00%'

    sheet.cell(row=next_row, column=len(command) + 3 - 2).value = efficiency
    sheet.cell(row=next_row, column=len(command) + 3 - 2).number_format = '0.00%'

    is_match, m_value, n_value, k_value, lda_value, ldb_value, ldc_value = check_mnkabc_match(
        command[8], command[9], command[10], command[11], command[12], command[13])

    if is_match == None:
        sheet.cell(row=next_row, column=len(command) + 10 - 2, value="Error")
    else:
        sheet.cell(row=next_row, column=len(command) + 10 - 2, value="OK")
        if is_match == False:
            sheet.cell(row=next_row, column=len(command) + 4 - 2, value=m_value)
            sheet.cell(row=next_row, column=len(command) + 5 - 2, value=n_value)
            sheet.cell(row=next_row, column=len(command) + 6 - 2, value=k_value)
            sheet.cell(row=next_row, column=len(command) + 7 - 2, value=lda_value)
            sheet.cell(row=next_row, column=len(command) + 8 - 2, value=ldb_value)
            sheet.cell(row=next_row, column=len(command) + 9 - 2, value=ldc_value)
            
    sheet.freeze_panes = 'A2'

    # Save the Excel workbook
    workbook.save('report.xlsx')

def get_type_dict():
    type_dict = {
        "R_16F": 232.7,
        "R_16F_C_32": 232.7,
        "R_32F": 29.1,
        "R_64F": 29.1,
        "R_16B": 232.7
    }
    return type_dict

def get_mnkabc_list():
    # Read the content of mnk.log
    with open('mnk.log', 'r') as file:
        content = file.read()

    # Use regular expressions to extract m, n, and k values from each line
    # pattern = r'(\d+),\s*(\d+),\s*(\d+)'
    pattern = r'(\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+)'

    matches = re.findall(pattern, content)

    # Store the extracted m, n, and k values in a list of tuples
    mnkabc_values = [(int(m), int(n), int(k), int(a), int(b), int(c)) for m, n, k, a, b, c in matches]


    # Print the extracted values
    #for m, n, k in mnk_values:
    #    print(f"m = {m}, n = {n}, k = {k}")
    return mnkabc_values

def get_gpu_clock(max_clock, stop_event):
    # clocks = []
    try:
        while not stop_event.is_set():
            output = subprocess.check_output(['rocm-smi', '--showgpuclocks']).decode('utf-8')
            lines = output.split('\n')
            for line in lines:
                if 'sclk clock level' in line:
                    clock_level = int(line.split('(')[-1].split('Mhz')[0])
                    print(f"GPU Clock: {clock_level} MHz")
                    # clocks.append(clock_level)

                    # max_clock[0] = max(clocks)  # Update max_clock if a new maximum is found
                    # print("cocotion test = ", clock_level)
                    # max_clock = clock_level

                    max_clock[0] = max(max_clock[0], clock_level)
                    # print("cocotion test @@ max_clock = ", max_clock)

                    time.sleep(1)  # Wait for 1 second before capturing the next GPU clock
    except subprocess.CalledProcessError as e:
        print(f"Error running rocm-smi: {e}")

try:
    # Attempt to remove the report.xlsx file
    os.remove('report.xlsx')
    print("Successfully removed report.xlsx")
except FileNotFoundError:
    print("report.xlsx does not exist. Skipping removal.")


print("QQQQQQQQQQQQQQ")

base_gfreq = 1.42
peak_tflops = 24.6
type_dict = get_type_dict()
mnkabc_values = get_mnkabc_list()

print(mnkabc_values)

# Create a list to store the maximum clock frequency
max_clock = [0]

# Define the command with parameters
command = ['python3', 'qgemm.py', 'R_32F', 'R_32F', 'R_32F', 'R_32F', 'OP_N', 'OP_T',
           '8640', '8640', '8640', '8640', '8640', '8640', '72', '30']

# ./gemm R_16F R_16F R_32F R_32F OP_N OP_T 8640 8640 8640 8640 8640 8640 72 30 ex
for type_key in type_dict:
    if type_key == "R_16F":
        command = ['python3', 'qgemm.py', 'R_16F', 'R_16F', 'R_16F', 'R_16F', 'OP_N', 'OP_T',
            '8640', '8640', '8640', '8640', '8640', '8640', '72', '30', 'ex']
    elif type_key == "R_16B":
        command = ['python3', 'qgemm.py', 'R_16B', 'R_16B', 'R_32F', 'R_32F', 'OP_N', 'OP_T',
            '8640', '8640', '8640', '8640', '8640', '8640', '72', '30', 'ex']
    elif type_key == "R_16F_C_32":
        command = ['python3', 'qgemm.py', 'R_16F', 'R_16F', 'R_32F', 'R_32F', 'OP_N', 'OP_T',
            '8640', '8640', '8640', '8640', '8640', '8640', '72', '30', 'ex']

    if type_key != "R_16F_C_32" and type_key != "R_16B":
        command[2] = type_key
        command[3] = type_key
        command[4] = type_key
        command[5] = type_key
    
    # if type_key == "R_16B":
    #     command[4] = "R_32F"
    #     command[5] = "R_32F"


    for m, n, k, a, b, c in mnkabc_values:
        command[8] = str(m)
        command[9] = str(n)
        command[10] = str(k)
        command[11] = str(a)
        command[12] = str(b)
        command[13] = str(c)
        command[14] = "72"

        print(command)


        while int(command[14]) >= 1:
            stop_event = threading.Event()
            # Create a thread to run the get_gpu_clock function in the background
            clock_thread = threading.Thread(target=get_gpu_clock, args=(max_clock, stop_event))
            clock_thread.daemon = True  # Set the thread as a daemon so it will exit when the main program exits
            clock_thread.start()

            print("cocotion test start exe")
            execute_program(command)
            print("cocotion test after exe")

            stop_event.set()
            clock_thread.join()
            # clock_thread.terminate()
            print(f"Maximum GPU Clock: {max_clock[0]} MHz")
            real_gpu_gfreq = max_clock[0] / 1000
            max_clock = [0]

            tflops = get_TFLOPS()
            if tflops != None and tflops > 0:
                break
            command[14] = str(int(command[14]) // 2)

        produce_report(command, tflops, (type_dict[type_key] / base_gfreq) * real_gpu_gfreq)


# execute_program(command)
# tflops = get_TFLOPS()

# produce_report(command, tflops, peak_tflops)
